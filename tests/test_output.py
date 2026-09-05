from __future__ import annotations

import os
import tempfile

import openpyxl

from engine.allocation import SkuAllocation
from engine.dos_difc import DIFCResult, DIFCRow
from engine.engine_result import EngineResult
from engine.fallback import FallbackDecisions
from engine.reconciliation import ReconciledResult
from output import comparison_table_sheet, excel_writer, weekly_plan_sheet


def _sample_engine_result(fallback_applied: bool) -> EngineResult:
    alloc = SkuAllocation(
        plant_line="PlantA_Line1", period=1, link_code="L1", sku="L1", priority=1.0,
        current_fin=100.0, carryover_fin_in=0.0, wk1=100.0,
        plant="PlantA", line="Line1", brand="BrandA", link_desc="Product A",
        month_key="Feb-26", opening_dos=20.0, target_dos=30.0, moq_days=5.0,
    )
    reconciled = ReconciledResult(rows=[alloc])
    difc = DIFCResult(rows=[DIFCRow(plant_line="PlantA_Line1", period=1, link_code="L1",
                                     sku="L1", closing_by_week={"wk1": 25.0}, approximated=False,
                                     plant="PlantA", line="Line1", brand="BrandA",
                                     link_desc="Product A", month_key="Feb-26", opening_dos=20.0)])
    fb = FallbackDecisions(
        use_default_calendar=fallback_applied,
        messages=["Calendar not supplied — capacity is a coarse approximation."] if fallback_applied else [],
    )
    return EngineResult(reconciled=reconciled, difc=difc, fallback=fb, capacity_messages=[])


def test_workbook_has_all_four_tabs_with_expected_headers():
    result = _sample_engine_result(fallback_applied=False)
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "out.xlsx")
        excel_writer.write(result, path)
        wb = openpyxl.load_workbook(path)
        assert set(wb.sheetnames) == {
            "Weekly Plan", "Comparison Table", "Weekly DIFC Summary", "Assumption Applied",
        }

        weekly_plan = wb["Weekly Plan"]
        headers = [c.value for c in next(weekly_plan.iter_rows(max_row=1))]
        assert "TOTAL PRODUCED" in headers and "CARRYOVER_MPLUS1" in headers
        assert "SKU" not in headers  # Linkcode-level only, per client decision -- no SKU column anywhere


def test_weekly_plan_has_business_context_columns_matching_sample_format():
    result = _sample_engine_result(fallback_applied=False)
    df = weekly_plan_sheet.build_dataframe(result)
    for col in (
        "Plant", "Line", "Linkcode", "Brand", "Link Desc Description", "Month", "Period",
        "Opening DOS", "Target DOS", "Priority", "MOQ", "Notes",
    ):
        assert col in df.columns
    row = df.iloc[0]
    assert row["Plant"] == "PlantA"
    assert row["Line"] == "Line1"
    assert row["Linkcode"] == "L1"
    assert row["Brand"] == "BrandA"
    assert row["Link Desc Description"] == "Product A"
    assert row["Month"] == "Feb-26"
    assert row["Opening DOS"] == 20.0
    assert row["Target DOS"] == 30.0
    assert row["MOQ"] == 5.0


def test_weekly_plan_notes_column_joins_row_assumptions():
    alloc = SkuAllocation(
        plant_line="P_L", period=1, link_code="L1", sku="L1", priority=1.0,
        current_fin=100.0, carryover_fin_in=0.0, wk1=100.0,
        assumptions=["Priority defaulted to file order (no RCCP match)."],
    )
    result = EngineResult(reconciled=ReconciledResult(rows=[alloc]), difc=DIFCResult(rows=[]),
                          fallback=FallbackDecisions(), capacity_messages=[])
    df = weekly_plan_sheet.build_dataframe(result)
    assert df.iloc[0]["Notes"] == "Priority defaulted to file order (no RCCP match)."


def test_header_row_is_styled_purple_with_white_bold_text():
    result = _sample_engine_result(fallback_applied=False)
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "out.xlsx")
        excel_writer.write(result, path)
        wb = openpyxl.load_workbook(path)
        for sheet_name in ("Weekly Plan", "Comparison Table", "Weekly DIFC Summary", "Assumption Applied"):
            cell = wb[sheet_name]["A1"]
            assert cell.fill.fgColor.rgb == "FF4F2170"
            assert cell.font.bold is True
            assert cell.font.color.rgb == "FFFFFFFF"


def test_assumptions_tab_empty_when_no_fallback_applied():
    result = _sample_engine_result(fallback_applied=False)
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "out.xlsx")
        excel_writer.write(result, path)
        wb = openpyxl.load_workbook(path)
        sheet = wb["Assumption Applied"]
        assert sheet.max_row == 1  # header only, no data rows


def test_assumptions_tab_populated_when_fallback_applied():
    result = _sample_engine_result(fallback_applied=True)
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "out.xlsx")
        excel_writer.write(result, path)
        wb = openpyxl.load_workbook(path)
        sheet = wb["Assumption Applied"]
        assert sheet.max_row > 1
        assert "Calendar not supplied" in sheet.cell(row=2, column=1).value


def test_comparison_table_has_moq_case_column():
    # The spec lists "MOQ compliance flags" for COMPARISON_TABLE; it surfaces as
    # the per-SKU Run 1 case (A/B/C/D or "No MOQ").
    a = SkuAllocation(plant_line="P_L1", period=1, link_code="L1", sku="L1", priority=1.0,
                      current_fin=100.0, carryover_fin_in=0.0, wk1=100.0, moq_case="D")
    b = SkuAllocation(plant_line="P_L1", period=1, link_code="L2", sku="L2", priority=2.0,
                      current_fin=50.0, carryover_fin_in=0.0, wk1=50.0, moq_case="No MOQ")
    result = EngineResult(reconciled=ReconciledResult(rows=[a, b]), difc=DIFCResult(rows=[]),
                          fallback=FallbackDecisions(), capacity_messages=[])
    df = comparison_table_sheet.build_dataframe(result)
    assert "CASE" in df.columns
    assert df.set_index("Linkcode")["CASE"].to_dict() == {"L1": "D", "L2": "No MOQ"}
