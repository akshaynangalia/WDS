"""
End-to-end test of the full call sequence in Development Planning Document,
Section 2.3 -- parsers through to a downloadable workbook -- using small,
CONSTRUCTED, date-aligned workbooks (not the client's real sample files,
which don't share a common period range with each other; see the Risk
Register). This is the test that proves the whole pipeline is actually wired
together correctly, headless, with no Dash/UI involved (Rule 3).
"""
from __future__ import annotations

import os
import tempfile

import openpyxl
import pandas as pd

from orchestration.models import RunParams
from orchestration.run_manager import execute_run


def _build_mps_input(path: str):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame({
            "Link Code": [111111], "SKU": [111111], "Brand": ["TestBrand"],
            "Link Desc Description": ["Test Product"],
        }).to_excel(writer, sheet_name="SKU Master", index=False)

        pd.DataFrame({"Link Code": [111111], 1: [620.0]}).to_excel(
            writer, sheet_name="2.Demand Input", index=False)

        pd.DataFrame({"Key": [pd.Timestamp("2026-06-01")], "Period": [1]}).to_excel(
            writer, sheet_name="Period Calendar Matrix", index=False)

        pd.DataFrame({
            "Link Code": [111111], "Period": [1], "Plant": ["TestPlant"], "Line": ["Line1"],
            "GE%": [1.0], "SOC": [24.0],
        }).to_excel(writer, sheet_name="4.SOC Sheet & Flag", index=False)


def _build_mps_output(path: str):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame({
            "Period": [1], "SKU": [111111], "Brand": ["TestBrand"], "Link Code": [111111],
            "Link Desc Description": ["Test Product"], "O/S": [10], "DOS": [20],
            "TestPlant_Line1": [300.0],
        }).to_excel(writer, sheet_name="SKU Line Loading 1", index=False)

        pd.DataFrame({
            "Link Code": [111111], "Brand": ["TestBrand"], "Link Desc Description": ["Test Product"],
            1: [20.0], "Min": [20.0], "Max": [20.0], "Avg": [20.0], "Avg_min_dos_target": [20.0],
        }).to_excel(writer, sheet_name="Linkcode_DIFC", index=False)


def test_full_pipeline_headless_produces_valid_workbook():
    with tempfile.TemporaryDirectory() as tmp:
        mps_input_path = os.path.join(tmp, "mps_input.xlsx")
        mps_output_path = os.path.join(tmp, "mps_output.xlsx")
        _build_mps_input(mps_input_path)
        _build_mps_output(mps_output_path)

        params = RunParams(start_period=1, end_period=1)
        result = execute_run(mps_input_path, mps_output_path, None, params, output_dir=tmp)

        assert result.status.value == "degraded"  # Manual Input absent -> fallback path
        assert result.output_path and os.path.exists(result.output_path)
        assert any("Priority not supplied" in m for m in result.assumption_messages)
        assert any("Calendar not supplied" in m for m in result.assumption_messages)

        wb = openpyxl.load_workbook(result.output_path)
        weekly_plan = wb["Weekly Plan"]
        total_col = [c.value for c in next(weekly_plan.iter_rows(max_row=1))].index("TOTAL PRODUCED") + 1
        total_value = weekly_plan.cell(row=2, column=total_col).value
        assert round(total_value, 1) == 300.0  # must equal FIN -- the whole point of reconciliation


def test_fails_cleanly_when_mps_output_missing():
    with tempfile.TemporaryDirectory() as tmp:
        mps_input_path = os.path.join(tmp, "mps_input.xlsx")
        _build_mps_input(mps_input_path)
        params = RunParams(start_period=1, end_period=1)
        result = execute_run(mps_input_path, mps_input_path, None, params, output_dir=tmp)
        # passing mps_input twice means MPS Output's required sheets are missing
        assert result.status.value == "failed"
        assert result.output_path is None
        assert result.errors
