"""Shared header styling for every output tab, matching the client-supplied
sample workbook ("Sample Output Format.xlsx"): a purple fill with bold white
centered text and a thin border, applied to row 1 of each sheet.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="FF4F2170", end_color="FF4F2170", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
HEADER_BORDER = Border(*(Side(style="thin") for _ in range(4)))
HEADER_ROW_HEIGHT = 19.8


def style_header_row(ws: Worksheet, ncols: int) -> None:
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER
